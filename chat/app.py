#!/usr/bin/env python3
"""
Simple Chat Interface for Ollama Models
"""

from flask import Flask, render_template, request, jsonify, Response, send_file, after_this_request
import requests
import json
import os
import pandas as pd
from werkzeug.utils import secure_filename
import tempfile
from openpyxl import load_workbook

app = Flask(__name__)
OLLAMA_API = 'http://ollama:11434'
MODELS_DIR = '/models'

# Converter configuration
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
app.config['OUTPUT_FOLDER'] = '/data/training'

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/converter')
def converter_page():
    return render_template('converter.html')

@app.route('/api/converter/convert', methods=['POST'])
def convert():
    temp_jsonl = None
    try:
        # Validate file upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Use .xlsx, .xls, or .csv'}), 400
        
        # Get column mappings and save option
        instruction_col = request.form.get('instruction_col', 'instruction')
        output_col = request.form.get('output_col', 'output')
        save_to_server = request.form.get('save_to_server', 'false').lower() == 'true'
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Read file
        if filename.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
        
        # Validate columns
        if instruction_col not in df.columns:
            os.remove(filepath)
            return jsonify({'error': f'Column "{instruction_col}" not found in file'}), 400
        if output_col not in df.columns:
            os.remove(filepath)
            return jsonify({'error': f'Column "{output_col}" not found in file'}), 400
        
        # Convert to JSONL
        output_filename = filename.rsplit('.', 1)[0] + '.jsonl'
        
        # Create temporary JSONL file
        temp_jsonl = tempfile.NamedTemporaryFile(mode='w', suffix='.jsonl', delete=False, encoding='utf-8')
        rows_converted = 0
        
        for _, row in df.iterrows():
            if pd.notna(row[instruction_col]) and pd.notna(row[output_col]):
                json_obj = {
                    'instruction': str(row[instruction_col]).strip(),
                    'output': str(row[output_col]).strip()
                }
                temp_jsonl.write(json.dumps(json_obj, ensure_ascii=False) + '\n')
                rows_converted += 1
        
        temp_jsonl.close()
        
        # Optionally save to server
        if save_to_server:
            output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
            os.rename(temp_jsonl.name, output_path)
            os.remove(filepath)
            return jsonify({
                'success': True,
                'message': f'Converted and saved to {output_filename}',
                'output_file': output_filename,
                'rows_converted': rows_converted
            })
        
        # Clean up uploaded file
        os.remove(filepath)
        
        # Return file as download with cleanup callback
        @after_this_request
        def remove_file(response):
            try:
                os.remove(temp_jsonl.name)
            except Exception as error:
                app.logger.error(f"Error removing temp file: {error}")
            return response
        
        return send_file(
            temp_jsonl.name,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/jsonl'
        )
    
    except Exception as e:
        app.logger.error(f"Conversion error: {str(e)}")
        if temp_jsonl and hasattr(temp_jsonl, 'name') and os.path.exists(temp_jsonl.name):
            try:
                os.remove(temp_jsonl.name)
            except:
                pass
        return jsonify({'error': str(e)}), 500

@app.route('/api/converter/preview', methods=['POST'])
def preview():
    filepath = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'preview_' + filename)
        file.save(filepath)
        
        # Read file
        if filename.endswith('.csv'):
            df = pd.read_csv(filepath, nrows=5)
            # Clean up
            if filepath and os.path.exists(filepath):
                os.remove(filepath)
            
            return jsonify({
                'columns': df.columns.tolist(),
                'preview': df.head().to_dict('records')
            })
        else:
            # Read Excel using openpyxl in read-only mode for performance
            wb = None
            try:
                wb = load_workbook(filename=filepath, read_only=True, data_only=True)
                ws = wb.active
                
                columns = []
                data = []
                
                # Iterate rows (generator) to avoid loading full file
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        columns = list(row)
                    elif i <= 5: # Limit to 5 data rows
                        row_data = {}
                        for col_idx, cell_value in enumerate(row):
                            if col_idx < len(columns):
                                col_name = columns[col_idx]
                                row_data[col_name] = cell_value
                        data.append(row_data)
                    else:
                        break
            finally:
                if wb:
                    wb.close()
            
            # Clean up
            if filepath and os.path.exists(filepath):
                os.remove(filepath)

            return jsonify({
                'columns': columns,
                'preview': data
            })
    
    except Exception as e:
        app.logger.error(f"Preview error: {str(e)}")
        # Clean up on error
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except:
                pass
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    return render_template('chat.html')

@app.route('/api/models')
def get_models():
    try:
        response = requests.get(f'{OLLAMA_API}/api/tags')
        return jsonify(response.json())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/modelfiles')
def get_modelfiles():
    modelfiles = []
    try:
        # Walk through models directory to find Modelfiles
        for root, dirs, files in os.walk(MODELS_DIR):
            if 'Modelfile' in files:
                full_path = os.path.join(root, 'Modelfile')
                # Create a relative path for display, removing /models/ prefix
                rel_path = os.path.relpath(full_path, MODELS_DIR)
                modelfiles.append({
                    'path': full_path,
                    'name': rel_path
                })
        
        # Sort by name
        modelfiles.sort(key=lambda x: x['name'])
        return jsonify({'modelfiles': modelfiles})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pull-model', methods=['POST'])
def pull_model():
    try:
        data = request.json
        model_name = data.get('name')

        if not model_name:
            return jsonify({'error': 'Missing model name'}), 400

        def generate():
            try:
                response = requests.post(
                    f'{OLLAMA_API}/api/pull',
                    json={'name': model_name},
                    stream=True
                )

                for line in response.iter_lines():
                    if line:
                        yield line.decode('utf-8') + '\n'
            except Exception as e:
                yield json.dumps({'error': str(e)}) + '\n'

        return Response(generate(), mimetype='text/event-stream')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/create-model', methods=['POST'])
def create_model():
    import subprocess
    import re
    try:
        data = request.json
        model_name = data.get('name')
        modelfile_path = data.get('path')

        if not model_name or not modelfile_path:
            return jsonify({'error': 'Missing name or path'}), 400

        # Check if Modelfile exists
        if not os.path.exists(modelfile_path):
             return jsonify({'error': 'Modelfile not found'}), 404

        # Stream response from Ollama CLI
        def generate():
            try:
                yield json.dumps({'status': f'Creating model {model_name}...'}) + '\n'

                # Use ollama CLI command via docker exec directly
                # The chat container can exec into the ollama container
                process = subprocess.Popen(
                    ['docker', 'exec', 'ollama', 'ollama', 'create', model_name, '-f', modelfile_path],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1
                )

                # Stream the output
                for line in process.stdout:
                    line = line.strip()
                    if line:
                        # Remove ANSI escape codes
                        line = re.sub(r'\x1b\[[0-9;]*[a-zA-Z?]', '', line)
                        yield json.dumps({'status': line}) + '\n'

                # Wait for process to complete
                process.wait()

                if process.returncode == 0:
                    yield json.dumps({'status': 'success'}) + '\n'
                else:
                    yield json.dumps({'error': f'Process exited with code {process.returncode}'}) + '\n'

            except Exception as e:
                yield json.dumps({'error': str(e)}) + '\n'

        return Response(generate(), mimetype='text/event-stream')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/modelfile', methods=['GET'])
def get_modelfile():
    try:
        path = request.args.get('path')
        if not path or not os.path.exists(path):
            return jsonify({'error': 'Modelfile not found'}), 404

        with open(path, 'r') as f:
            content = f.read()

        return jsonify({'content': content, 'path': path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/modelfile', methods=['POST'])
def create_modelfile():
    try:
        data = request.json
        name = data.get('name')
        content = data.get('content', '')

        if not name:
            return jsonify({'error': 'Missing name'}), 400

        # Create in models/custom directory
        modelfile_dir = os.path.join(MODELS_DIR, 'custom', name)
        os.makedirs(modelfile_dir, exist_ok=True)

        modelfile_path = os.path.join(modelfile_dir, 'Modelfile')

        if os.path.exists(modelfile_path):
            return jsonify({'error': 'Modelfile already exists'}), 409

        with open(modelfile_path, 'w') as f:
            f.write(content)

        return jsonify({'success': True, 'path': modelfile_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/modelfile', methods=['PUT'])
def update_modelfile():
    try:
        data = request.json
        path = data.get('path')
        content = data.get('content')

        if not path or content is None:
            return jsonify({'error': 'Missing path or content'}), 400

        if not os.path.exists(path):
            return jsonify({'error': 'Modelfile not found'}), 404

        with open(path, 'w') as f:
            f.write(content)

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/modelfile', methods=['DELETE'])
def delete_modelfile():
    try:
        path = request.args.get('path')

        if not path or not os.path.exists(path):
            return jsonify({'error': 'Modelfile not found'}), 404

        # Only allow deletion from custom directory for safety
        if '/custom/' not in path:
            return jsonify({'error': 'Can only delete custom Modelfiles'}), 403

        os.remove(path)

        # Try to remove parent directory if empty
        parent_dir = os.path.dirname(path)
        try:
            if not os.listdir(parent_dir):
                os.rmdir(parent_dir)
        except:
            pass

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete-model', methods=['DELETE'])
def delete_model():
    try:
        model_name = request.args.get('name')

        if not model_name:
            return jsonify({'error': 'Missing model name'}), 400

        response = requests.delete(
            f'{OLLAMA_API}/api/delete',
            json={'name': model_name}
        )

        if response.status_code == 200:
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Failed to delete model'}), response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        data = request.json
        model = data.get('model', 'llama3.2:1b')
        message = data.get('message', '')

        # Stream response from Ollama
        def generate():
            response = requests.post(
                f'{OLLAMA_API}/api/generate',
                json={
                    'model': model,
                    'prompt': message,
                    'stream': True
                },
                stream=True
            )

            for line in response.iter_lines():
                if line:
                    yield line.decode('utf-8') + '\n'

        return Response(generate(), mimetype='text/event-stream')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health():
    return jsonify({'status': 'healthy'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=True)
